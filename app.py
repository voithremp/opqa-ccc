import streamlit as st
import pandas as pd
import re
import os
from datetime import datetime
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ====== CLEAR FUNCTIONS ======

def process_bet_file(file, table_dict, column_name):
    try:
        df = pd.read_csv(file, encoding='utf-16', header=0, sep=None, engine='python')
    except Exception as e:
        st.error(f"Failed to read file: {file}\n\n{e}")
        return

    for _, row in df.iterrows():
        if len(row) < 3:
            continue
        ids_field = str(row.iloc[1])
        setting = str(row.iloc[2]).strip()
        for table_id in table_dict:
            if table_id in ids_field:
                if table_dict[table_id][column_name] == '':
                    table_dict[table_id][column_name] = setting

def run_clear(id_file, large_file, medium_file, small_file, xsmall_file):
    try:
        id_df = pd.read_csv(id_file, encoding='utf-8', header=0)
    except Exception as e:
        st.error(f"Failed to read Table ID file:\n\n{e}")
        return None

    table_dict = {}
    for val in id_df.iloc[:, 0]:
        table_id = str(val).strip()
        if re.fullmatch(r'[A-Za-z0-9-]{16}', table_id):
            table_dict[table_id] = {'tableID': table_id, 'large': '', 'medium': '', 'small': '', 'xsmall': ''}

    process_bet_file(large_file, table_dict, 'large')
    process_bet_file(medium_file, table_dict, 'medium')
    process_bet_file(small_file, table_dict, 'small')
    process_bet_file(xsmall_file, table_dict, 'xsmall')

    output_df = pd.DataFrame(table_dict.values())
    return output_df

# ====== COMPARE FUNCTIONS ======

def extract_first_parameters(text):
    params = {}
    if not isinstance(text, str):
        return params
    matches = re.findall(r'(\w[\w\-]*?)=([^\n]*)', text)
    for key, value in matches:
        if key not in params:
            params[key] = value.strip()
    return params

def extract_clean_parameters(text):
    params = {}
    if not isinstance(text, str):
        return params
    lines = text.splitlines()
    for line in lines:
        if '=' in line:
            key, value = line.split('=', 1)
            params[key.strip()] = value.strip()
    return params

def run_compare(file_a, file_b, threshold):
    df_a = pd.read_excel(file_a)
    df_b = pd.read_excel(file_b)

    a_data = {}
    for idx, row in df_a.iterrows():
        table_id = str(row.iloc[0]).strip()
        if table_id == 'nan':
            continue
        a_data[table_id] = {
            'Large': extract_first_parameters(row.iloc[1]),
            'Medium': extract_first_parameters(row.iloc[2]),
            'Small': extract_first_parameters(row.iloc[3]),
            'XSmall': extract_first_parameters(row.iloc[4]),
        }

    b_data = []
    b_table_id_count = {}

    for idx, row in df_b.iterrows():
        table_id = str(row.iloc[0]).strip()
        if table_id == 'nan':
            continue
        b_data.append({
            'TableID': table_id,
            'Large': extract_clean_parameters(row.iloc[1]),
            'Medium': extract_clean_parameters(row.iloc[2]),
            'Small': extract_clean_parameters(row.iloc[3]),
            'XSmall': extract_clean_parameters(row.iloc[4]),
        })
        b_table_id_count[table_id] = b_table_id_count.get(table_id, 0) + 1

    duplicate_ids = [tid for tid, count in b_table_id_count.items() if count > 1]
    if duplicate_ids:
        st.warning(f"⚠️ Duplicate Table IDs found and excluded: {duplicate_ids}")

    b_data = [item for item in b_data if item['TableID'] not in duplicate_ids]

    missing_in_a = []
    output_rows = []
    flagged_tables = {}

    for item in b_data:
        table_id = item['TableID']
        if table_id not in a_data:
            missing_in_a.append(table_id)
            continue

        row = {
            'TableID': table_id,
            'Wrong (Large)': '',
            'Wrong (Medium)': '',
            'Wrong (Small)': '',
            'Wrong (XSmall)': '',
            'Full Correct (Large)': '',
            'Full Correct (Medium)': '',
            'Full Correct (Small)': '',
            'Full Correct (XSmall)': '',
        }

        table_flag = []

        for size in ['Large', 'Medium', 'Small', 'XSmall']:
            a_params = a_data[table_id][size]
            b_params = item[size]
            wrongs = []
            corrects = []

            for param, b_value in b_params.items():
                if param not in a_params:
                    wrongs.append(f"{param}={b_value}")
                else:
                    a_value = a_params[param]
                    if a_value != b_value:
                        wrongs.append(f"{param}={b_value}")
                        corrects.append(f"{param}={a_value}")
                    else:
                        corrects.append(f"{param}={a_value}")

            if wrongs:
                row[f'Wrong ({size})'] = '\n'.join(wrongs)
                row[f'Full Correct ({size})'] = '\n'.join(corrects)

            if len(wrongs) > threshold:
                table_flag.append(size)

        if table_flag:
            flagged_tables[table_id] = table_flag

        output_rows.append(row)

    if missing_in_a:
        st.warning(f"⚠️ Table IDs found in B but not in A: {missing_in_a}")

    output_df = pd.DataFrame(output_rows)

    # Create Excel with highlights
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        output_df.to_excel(writer, index=False, sheet_name='Comparison')

        wb = writer.book
        ws = writer.sheets['Comparison']

        red_fill = PatternFill(start_color='FFF4CCCC', end_color='FFF4CCCC', fill_type='solid')
        bold_font = Font(bold=True)

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 17

        ws.freeze_panes = ws['A2']

        tableid_col = 1
        wrong_cols = {'Large': 2, 'Medium': 3, 'Small': 4, 'XSmall': 5}

        for row_idx in range(2, ws.max_row + 1):
            table_id = ws.cell(row=row_idx, column=tableid_col).value
            if table_id in flagged_tables:
                ws.cell(row=row_idx, column=tableid_col).fill = red_fill
                ws.cell(row=row_idx, column=tableid_col).font = bold_font
                for size in flagged_tables[table_id]:
                    ws.cell(row=row_idx, column=wrong_cols[size]).fill = red_fill

    buffer.seek(0)
    return buffer

# ====== STREAMLIT APP ======

st.set_page_config(page_title="Table Config Processor", layout="wide")
st.title("📋 Table Config Processor (Clear + Compare)")

tabs = st.tabs(["🧹 Clear & Match", "🆚 Compare"])

# --- Clear Tab
with tabs[0]:
    st.header("🧹 Clear and Match")

    id_file = st.file_uploader("Upload Table ID File (UTF-8 CSV)", type="csv")
    large_file = st.file_uploader("Upload Large Bet Limit File (UTF-16 CSV)", type="csv")
    medium_file = st.file_uploader("Upload Medium Bet Limit File (UTF-16 CSV)", type="csv")
    small_file = st.file_uploader("Upload Small Bet Limit File (UTF-16 CSV)", type="csv")
    xsmall_file = st.file_uploader("Upload XSmall Bet Limit File (UTF-16 CSV)", type="csv")

    if st.button("🚀 Run Clear & Match"):
        if not all([id_file, large_file, medium_file, small_file, xsmall_file]):
            st.error("Please upload ALL 5 files.")
        else:
            output_df = run_clear(id_file, large_file, medium_file, small_file, xsmall_file)
            if output_df is not None:
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    output_df.to_excel(writer, index=False, sheet_name='ClearResult')
                buffer.seek(0)

                st.success("✅ Clear & Match completed!")
                st.download_button(
                    label="Download Clear Result Excel",
                    data=buffer,
                    file_name="clear_result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

# --- Compare Tab
with tabs[1]:
    st.header("🆚 Compare Two Files")

    file_a = st.file_uploader("Upload File A (OPQA Template, Excel)", type=["xlsx"], key="file_a")
    file_b = st.file_uploader("Upload File B (ETI External, Excel)", type=["xlsx"], key="file_b")
    threshold = st.number_input("Wrong Threshold (per size)", min_value=1, max_value=20, value=5)

    if st.button("🚀 Run Comparison"):
        if not all([file_a, file_b]):
            st.error("Please upload both File A and File B.")
        else:
            compare_result = run_compare(file_a, file_b, threshold)
            if compare_result:
                st.success("✅ Comparison completed!")
                st.download_button(
                    label="Download Comparison Result Excel",
                    data=compare_result,
                    file_name="comparison_result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
